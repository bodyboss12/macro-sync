"""Tests for the Excel pivot exporter."""
from __future__ import annotations

import io
from datetime import date

import openpyxl
import pytest

from macro_sync.schema import DailySummary, NutritionEntry
from macro_sync.exporters.excel_pivot_exporter import (
    entries_to_pivot_excel_bytes,
    entries_to_pivot_workbook,
)


@pytest.fixture
def sample_entry() -> NutritionEntry:
    return NutritionEntry(
        date=date(2024, 1, 15),
        source="mfp",
        calories=500.0,
        protein_g=30.0,
        carbs_g=60.0,
        fat_g=15.0,
        fiber_g=5.0,
        sugar_g=10.0,
        sodium_mg=800.0,
    )


@pytest.fixture
def sample_summary() -> DailySummary:
    return DailySummary(
        date=date(2024, 1, 15),
        total_calories=500.0,
        total_protein_g=30.0,
        total_carbs_g=60.0,
        total_fat_g=15.0,
        total_fiber_g=5.0,
        entry_count=1,
    )


def test_pivot_excel_bytes_returns_bytes(sample_entry, sample_summary):
    result = entries_to_pivot_excel_bytes([sample_entry], [sample_summary])
    assert isinstance(result, bytes)


def test_pivot_excel_bytes_nonempty(sample_entry, sample_summary):
    result = entries_to_pivot_excel_bytes([sample_entry], [sample_summary])
    assert len(result) > 0


def test_pivot_workbook_has_two_sheets(sample_entry, sample_summary):
    wb = entries_to_pivot_workbook([sample_entry], [sample_summary])
    assert len(wb.sheetnames) == 2


def test_pivot_workbook_sheet_names(sample_entry, sample_summary):
    wb = entries_to_pivot_workbook([sample_entry], [sample_summary])
    assert "Entries" in wb.sheetnames
    assert "Daily Summary" in wb.sheetnames


def test_pivot_entries_sheet_header(sample_entry, sample_summary):
    wb = entries_to_pivot_workbook([sample_entry], [sample_summary])
    ws = wb["Entries"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
    assert "date" in headers
    assert "calories" in headers
    assert "source" in headers


def test_pivot_entries_data_row(sample_entry, sample_summary):
    wb = entries_to_pivot_workbook([sample_entry], [sample_summary])
    ws = wb["Entries"]
    assert ws.cell(row=2, column=1).value == "2024-01-15"
    assert ws.cell(row=2, column=3).value == 500.0


def test_pivot_summary_sheet_header(sample_entry, sample_summary):
    wb = entries_to_pivot_workbook([sample_entry], [sample_summary])
    ws = wb["Daily Summary"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
    assert "date" in headers
    assert "total_calories" in headers
    assert "entry_count" in headers


def test_pivot_summary_data_row(sample_entry, sample_summary):
    wb = entries_to_pivot_workbook([sample_entry], [sample_summary])
    ws = wb["Daily Summary"]
    assert ws.cell(row=2, column=1).value == "2024-01-15"
    assert ws.cell(row=2, column=2).value == 500.0
    assert ws.cell(row=2, column=7).value == 1


def test_pivot_excel_magic_bytes(sample_entry, sample_summary):
    result = entries_to_pivot_excel_bytes([sample_entry], [sample_summary])
    # xlsx files are ZIP archives starting with PK
    assert result[:2] == b"PK"
