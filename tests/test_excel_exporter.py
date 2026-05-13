"""Tests for the Excel exporter."""

from __future__ import annotations

import io
from datetime import date

import pytest

pytest.importorskip("openpyxl")
import openpyxl

from macro_sync.exporters.excel_exporter import (
    entries_to_excel_bytes,
    entries_to_workbook,
    summaries_to_excel_bytes,
    summaries_to_workbook,
)
from macro_sync.schema import DailySummary, NutritionEntry


@pytest.fixture
def sample_entry() -> NutritionEntry:
    return NutritionEntry(
        date=date(2024, 3, 15),
        source="myfitnesspal",
        food="Oatmeal",
        calories=300.0,
        protein=10.0,
        carbs=55.0,
        fat=5.0,
        fiber=4.0,
    )


@pytest.fixture
def sample_summary() -> DailySummary:
    return DailySummary(
        date=date(2024, 3, 15),
        total_calories=1800.0,
        total_protein=120.0,
        total_carbs=200.0,
        total_fat=60.0,
        total_fiber=25.0,
        entry_count=5,
    )


def test_entries_workbook_has_correct_sheet(sample_entry):
    wb = entries_to_workbook([sample_entry])
    assert "Entries" in wb.sheetnames


def test_entries_workbook_header_row(sample_entry):
    wb = entries_to_workbook([sample_entry])
    ws = wb["Entries"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
    assert headers == ["date", "source", "food", "calories", "protein", "carbs", "fat", "fiber"]


def test_entries_workbook_data_row(sample_entry):
    wb = entries_to_workbook([sample_entry])
    ws = wb["Entries"]
    assert ws.cell(row=2, column=1).value == "2024-03-15"
    assert ws.cell(row=2, column=2).value == "myfitnesspal"
    assert ws.cell(row=2, column=3).value == "Oatmeal"
    assert ws.cell(row=2, column=4).value == 300.0


def test_entries_workbook_row_count(sample_entry):
    wb = entries_to_workbook([sample_entry, sample_entry])
    ws = wb["Entries"]
    # 1 header + 2 data rows
    assert ws.max_row == 3


def test_entries_to_excel_bytes_is_valid_xlsx(sample_entry):
    raw = entries_to_excel_bytes([sample_entry])
    assert isinstance(raw, bytes)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert "Entries" in wb.sheetnames


def test_summaries_workbook_has_correct_sheet(sample_summary):
    wb = summaries_to_workbook([sample_summary])
    assert "Summaries" in wb.sheetnames


def test_summaries_workbook_header_row(sample_summary):
    wb = summaries_to_workbook([sample_summary])
    ws = wb["Summaries"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
    assert "date" in headers
    assert "total_calories" in headers
    assert "entry_count" in headers


def test_summaries_workbook_data_row(sample_summary):
    wb = summaries_to_workbook([sample_summary])
    ws = wb["Summaries"]
    assert ws.cell(row=2, column=1).value == "2024-03-15"
    assert ws.cell(row=2, column=2).value == 1800.0
    assert ws.cell(row=2, column=7).value == 5


def test_summaries_to_excel_bytes_is_valid_xlsx(sample_summary):
    raw = summaries_to_excel_bytes([sample_summary])
    assert isinstance(raw, bytes)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert "Summaries" in wb.sheetnames
