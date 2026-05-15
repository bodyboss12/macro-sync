"""Export nutrition data as Excel workbook with pivot-style summary sheet."""
from __future__ import annotations

from datetime import date
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from macro_sync.schema import DailySummary, NutritionEntry

_ENTRY_HEADERS = ["date", "source", "calories", "protein_g", "carbs_g", "fat_g", "fiber_g", "sugar_g", "sodium_mg"]
_SUMMARY_HEADERS = ["date", "total_calories", "total_protein_g", "total_carbs_g", "total_fat_g", "total_fiber_g", "entry_count"]
_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_headers(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 4, 12)


def _serialize(value) -> object:
    if isinstance(value, date):
        return value.isoformat()
    return value


def entries_to_pivot_workbook(entries: List[NutritionEntry], summaries: List[DailySummary]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    ws_entries = wb.active
    ws_entries.title = "Entries"
    _write_headers(ws_entries, _ENTRY_HEADERS)
    for row_idx, e in enumerate(entries, start=2):
        row = [_serialize(getattr(e, h, None)) for h in _ENTRY_HEADERS]
        for col_idx, val in enumerate(row, start=1):
            ws_entries.cell(row=row_idx, column=col_idx, value=val)

    ws_summary = wb.create_sheet(title="Daily Summary")
    _write_headers(ws_summary, _SUMMARY_HEADERS)
    for row_idx, s in enumerate(summaries, start=2):
        row = [_serialize(getattr(s, h, None)) for h in _SUMMARY_HEADERS]
        for col_idx, val in enumerate(row, start=1):
            ws_summary.cell(row=row_idx, column=col_idx, value=val)

    return wb


def entries_to_pivot_excel_bytes(entries: List[NutritionEntry], summaries: List[DailySummary]) -> bytes:
    import io
    wb = entries_to_pivot_workbook(entries, summaries)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
