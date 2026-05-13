"""Excel exporter for NutritionEntry and DailySummary objects."""

from __future__ import annotations

import io
from datetime import date
from typing import Iterable, Union

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError as exc:  # pragma: no cover
    raise ImportError(
        "openpyxl is required for Excel export. Install it with: pip install openpyxl"
    ) from exc

from macro_sync.schema import DailySummary, NutritionEntry

_ENTRY_HEADERS = ["date", "source", "food", "calories", "protein", "carbs", "fat", "fiber"]
_SUMMARY_HEADERS = ["date", "total_calories", "total_protein", "total_carbs", "total_fat", "total_fiber", "entry_count"]

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_headers(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def entries_to_workbook(entries: Iterable[NutritionEntry]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entries"
    _write_headers(ws, _ENTRY_HEADERS)
    for row_idx, entry in enumerate(entries, start=2):
        ws.cell(row=row_idx, column=1, value=entry.date.isoformat() if isinstance(entry.date, date) else entry.date)
        ws.cell(row=row_idx, column=2, value=entry.source)
        ws.cell(row=row_idx, column=3, value=entry.food)
        ws.cell(row=row_idx, column=4, value=entry.calories)
        ws.cell(row=row_idx, column=5, value=entry.protein)
        ws.cell(row=row_idx, column=6, value=entry.carbs)
        ws.cell(row=row_idx, column=7, value=entry.fat)
        ws.cell(row=row_idx, column=8, value=entry.fiber)
    return wb


def summaries_to_workbook(summaries: Iterable[DailySummary]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summaries"
    _write_headers(ws, _SUMMARY_HEADERS)
    for row_idx, summary in enumerate(summaries, start=2):
        ws.cell(row=row_idx, column=1, value=summary.date.isoformat() if isinstance(summary.date, date) else summary.date)
        ws.cell(row=row_idx, column=2, value=summary.total_calories)
        ws.cell(row=row_idx, column=3, value=summary.total_protein)
        ws.cell(row=row_idx, column=4, value=summary.total_carbs)
        ws.cell(row=row_idx, column=5, value=summary.total_fat)
        ws.cell(row=row_idx, column=6, value=summary.total_fiber)
        ws.cell(row=row_idx, column=7, value=summary.entry_count)
    return wb


def entries_to_excel_bytes(entries: Iterable[NutritionEntry]) -> bytes:
    wb = entries_to_workbook(entries)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def summaries_to_excel_bytes(summaries: Iterable[DailySummary]) -> bytes:
    wb = summaries_to_workbook(summaries)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
